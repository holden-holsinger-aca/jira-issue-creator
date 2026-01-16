from openpyxl import Workbook, load_workbook
from typing import List, Dict, Any


def extract_excel_info() -> List[Dict[str, Any]]:
    wb = load_workbook("tickets_to_create.xlsx", data_only=True)

    ws = wb.active

    tickets_to_add = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ticket = {"customfield_15377": {"value": "Review Workspace"}}
        ticket["issuetype"] = row[0]
        ticket["project"] = row[1]
        ticket["summary"] = row[2]
        tickets_to_add.append(ticket)

    return tickets_to_add


# need to create the parent epic or find the existing
